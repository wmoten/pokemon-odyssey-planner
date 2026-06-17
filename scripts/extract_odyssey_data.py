#!/usr/bin/env python3
"""Build the client-side Pokemon Odyssey data bundle from the v4.1.1 docs."""

from __future__ import annotations

import json
import os
import re
import urllib.error
import urllib.request
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
from pypdf import PdfReader
from PIL import Image

try:
    import pdfplumber
except Exception:  # pragma: no cover - the text extraction fallback still works.
    pdfplumber = None


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path.home() / "Downloads"

STATS_XLSX = Path(os.environ.get(
    "ODYSSEY_STATS_XLSX",
    DOWNLOADS / "Pokémon Stats, Learnset etc (v4.1.1).xlsx",
))
WILD_XLSX = Path(os.environ.get(
    "ODYSSEY_WILD_XLSX",
    DOWNLOADS / "Wild encounters, Items and TMs (v4.1.1).xlsx",
))
LEVELS_XLSX = Path(os.environ.get(
    "ODYSSEY_LEVELS_XLSX",
    DOWNLOADS / "Level Cap, Boss, Miniboss, Sea Map, Sidequests (v4.1.1).xlsx",
))
TEAMBUILDING_PDF = Path(os.environ.get(
    "ODYSSEY_TEAMBUILDING_PDF",
    DOWNLOADS / "Teambuilding options - v4.1.1.pdf",
))

OUT_JSON = ROOT / "src" / "data" / "pokemon-odyssey-data.json"
SPRITE_ASSET_DIR = ROOT / "src" / "assets" / "sprites"
SPRITE_ASSET_PREFIX = "src/assets/sprites"
STANDARD_ABILITY_CACHE = ROOT / "scripts" / "pokeapi-ability-definitions-cache.json"
POKEAPI_ABILITY_INDEX = "https://pokeapi.co/api/v2/ability?limit=10000"
HTTP_HEADERS = {"User-Agent": "pokemon-odyssey-team-planner/1.0"}
STANDARD_ABILITY_ALIASES = {
    "keeneyes": "keeneye",
}

STATS_SHEETS = ["#1-151", "#152-251", "#252-386", "4th Gen", "Paradox"]
POKEMON_START_COLS = [1, 4, 7]
ENCOUNTER_START_COLS = [1, 5, 9, 13]
STAT_NAMES = ["hp", "atk", "def", "spa", "spd", "spe"]
STONE_EVOLUTION_TERMS = [
    "Dawn Stone",
    "Dusk Stone",
    "Fire Stone",
    "Ice Stone",
    "Leaf Stone",
    "Link Stone",
    "Moon Stone",
    "Shiny Stone",
    "Sun Stone",
    "Thunderstone",
    "Water Stone",
]

METHOD_TOKENS = {
    "TALL GRASS",
    "HEADBUTT",
    "FISHING",
    "OLD ROD",
    "GOOD ROD",
    "SUPER ROD",
    "EVENT",
    "SPECIAL",
    "FLOOR",
    "CAVE",
    "ROCK SMASH",
    "GRASS",
    "SAND",
    "WATER",
    "SURF",
    "F.O.E.",
    "FOE",
}

HEADER_TOKENS = {
    "POKÉMON",
    "POKEMON",
    "LEVEL",
    "ENCOUNTER %",
    "NORMAL",
    "SHINY",
    "REFERENCE",
    "TYPE:",
    "ABILITY:",
    "EVOLUTION:",
    "MOVES",
}

EARLY_LOCATIONS = [
    "FIBERNIA",
    "ABANDONED LAB",
    "COASTAL ROAD",
    "SEASIDE GROTTO",
    "TALREGA",
    "VARLEY",
    "WATERFALL WOOD",
    "FIRST STRATUM",
    "SECOND STRATUM",
    "JAGGED MOUNTAIN",
    "MAPLE ISLAND",
    "DESERT OF GOLGONDA",
    "CAVE OF ATHARI",
    "NORTHERN ATOLL",
    "WONDER TRADE",
    "NORMAL ENCOUNTERS",
]

MID_LOCATIONS = [
    "THIRD STRATUM",
    "FOURTH STRATUM",
    "ARUNDEL",
    "CHARON",
    "GORENIL",
    "RADE OF CARCINO",
    "BRIGIT",
    "LOST WOODS",
    "AUBURN THICKET",
    "AZURE",
    "ETRIA GRAVEYARD",
    "PIRATE ISLAND",
    "AQUA RESORT",
    "AFTER THE FIRST STRATUM",
    "AFTER THE SECOND STRATUM",
    "AFTER THE THIRD STRATUM",
    "AFTER THE FOURTH STRATUM",
]

LATE_LOCATIONS = [
    "FIFTH STRATUM",
    "SIXTH STRATUM",
    "SEVENTH STRATUM",
    "EIGHTH STRATUM",
    "CAVE OF AGES",
    "FARAWAY",
    "POSTGAME",
    "ABYSSAL",
    "ACUITY",
    "AFTER THE FIFTH STRATUM",
    "AFTER THE SIXTH STRATUM",
    "AFTER THE SEVENTH STRATUM",
]

TIMING_ORDER = {"Starter": 0, "Early": 1, "Mid": 2, "Late": 3, "Postgame": 4, "Unknown": 9}

PROGRESSION_CHECKPOINTS = [
    {
        "id": "first-stratum",
        "sort": 1,
        "label": "First Stratum",
        "shortLabel": "First",
        "levelCap": 18,
        "description": "Available before clearing the First Stratum boss at Lv. 18.",
    },
    {
        "id": "second-stratum",
        "sort": 2,
        "label": "Second Stratum",
        "shortLabel": "Second",
        "levelCap": 30,
        "description": "Available after First Stratum and before clearing the Second Stratum boss at Lv. 30.",
    },
    {
        "id": "third-stratum",
        "sort": 3,
        "label": "Third Stratum",
        "shortLabel": "Third",
        "levelCap": 40,
        "description": "Available after Second Stratum and before clearing the Third Stratum boss at Lv. 40.",
    },
    {
        "id": "fourth-stratum",
        "sort": 4,
        "label": "Fourth Stratum",
        "shortLabel": "Fourth",
        "levelCap": 50,
        "description": "Available after Third Stratum and before clearing the Fourth Stratum boss at Lv. 50.",
    },
    {
        "id": "fifth-stratum",
        "sort": 5,
        "label": "Fifth Stratum",
        "shortLabel": "Fifth",
        "levelCap": 60,
        "description": "Available after Fourth Stratum and before clearing the Fifth Stratum boss at Lv. 60.",
    },
    {
        "id": "sixth-stratum",
        "sort": 6,
        "label": "Sixth Stratum",
        "shortLabel": "Sixth",
        "levelCap": 65,
        "description": "Available after Fifth Stratum and before clearing the Sixth Stratum boss at Lv. 65.",
    },
    {
        "id": "seventh-stratum",
        "sort": 7,
        "label": "Seventh Stratum",
        "shortLabel": "Seventh",
        "levelCap": 70,
        "description": "Available after Sixth Stratum and before clearing the Seventh Stratum boss at Lv. 70.",
    },
    {
        "id": "eighth-stratum",
        "sort": 8,
        "label": "Eighth Stratum",
        "shortLabel": "Eighth",
        "levelCap": 75,
        "description": "Available after Seventh Stratum and before clearing the Eighth Stratum boss at Lv. 75.",
    },
    {
        "id": "postgame",
        "sort": 9,
        "label": "Postgame",
        "shortLabel": "Postgame",
        "levelCap": None,
        "description": "Postgame-only documentation.",
    },
    {
        "id": "unknown",
        "sort": 99,
        "label": "Unknown",
        "shortLabel": "Unknown",
        "levelCap": None,
        "description": "No reliable progression source was parsed.",
    },
]

CHECKPOINT_BY_ID = {checkpoint["id"]: checkpoint for checkpoint in PROGRESSION_CHECKPOINTS}
CHECKPOINT_BY_SORT = {checkpoint["sort"]: checkpoint for checkpoint in PROGRESSION_CHECKPOINTS}

LOCATION_CHECKPOINT_RULES = [
    (8, ["EIGHTH STRATUM"]),
    (7, ["SEVENTH STRATUM", "PORCELAIN"]),
    (6, ["SIXTH STRATUM", "ACUITY", "LAKE OF RAGE", "MOUNT CHIMNEY"]),
    (5, ["FIFTH STRATUM", "AZURE", "LOST WOODS", "ETRIA GRAVEYARD", "CAVE OF AGES"]),
    (4, ["FOURTH STRATUM", "AUBURN THICKET"]),
    (3, ["THIRD STRATUM", "DESERT OF GOLGONDA", "CHARON", "BRIGIT", "MAPLE ISLAND"]),
    (2, ["SECOND STRATUM", "ARUNDEL", "YGGDRASIL", "CAVE OF ATHARI", "GORGE OF ZANADO"]),
    (
        1,
        [
            "FIBERNIA",
            "ABANDONED LAB",
            "COASTAL ROAD",
            "SEASIDE GROTTO",
            "TALREGA",
            "VARLEY",
            "WATERFALL WOOD",
            "NORTHERN ATOLL",
            "PIRATE ISLAND",
            "JAGGED MOUNTAIN",
            "FARAWAY ISLAND",
            "AQUA RESORT",
            "FIRST STRATUM",
            "NORMAL ENCOUNTERS",
            "WONDER TRADE",
        ],
    ),
]

LOTTERY_STONES = {"Fire Stone", "Water Stone", "Thunderstone", "Sun Stone"}
NAPIER_TREASURE_TIER_LEVELS = {0: 0, 1: 24, 2: 38, 3: 45, 4: 55, 5: 58}
ITEM_NAME_ALIASES = {
    "blackaugurite": "B. Augurite",
}

DIRECT_EVOLUTION_OVERRIDES = {
    "vileplume": ("gloom", "Leaf Stone"),
    "bellossom": ("gloom", "Sun Stone"),
    "poliwrath": ("poliwhirl", "Water Stone"),
    "politoed": ("poliwhirl", "Link Stone"),
    "slowbro": ("slowpoke", "LV.37"),
    "slowking": ("slowpoke", "Link Stone"),
    "kleavor": ("scyther", "B. Augurite"),
    "scizor": ("scyther", "Link Stone"),
    "hitmonlee": ("tyrogue", "LV.20"),
    "hitmonchan": ("tyrogue", "LV.20"),
    "hitmontop": ("tyrogue", "LV.20"),
    "vaporeon": ("eevee", "Water Stone"),
    "jolteon": ("eevee", "Thunderstone"),
    "flareon": ("eevee", "Fire Stone"),
    "espeon": ("eevee", "Happiness"),
    "umbreon": ("eevee", "Happiness"),
    "leafeon": ("eevee", "Musky Rock in First Stratum"),
    "glaceon": ("eevee", "Icy Stone in Arundel Woods"),
    "gardevoir": ("kirlia", "LV.30"),
    "gallade": ("kirlia", "Dawn Stone"),
    "glalie": ("snorunt", "LV.42"),
    "froslass": ("snorunt", "Dawn Stone"),
    "silcoon": ("wurmple", "LV.7"),
    "cascoon": ("wurmple", "LV.7"),
    "beautifly": ("silcoon", "LV.10"),
    "dustox": ("cascoon", "LV.10"),
    "sirfetchd": ("farfetchd-galar", "LV.35"),
}

SPECIAL_EVOLUTION_GATES = {
    "Musky Rock in First Stratum": {
        "checkpointId": "first-stratum",
        "source": "Teambuilding guide",
        "note": "Guide note says Leafeon evolves by interacting with the Musky Rock in the First Stratum.",
    },
    "Icy Stone in Arundel Woods": {
        "checkpointId": "second-stratum",
        "source": "Teambuilding guide",
        "note": "Guide note says Glaceon evolves by interacting with the Icy Stone in Arundel Woods.",
    },
}

NAME_ALIASES = {
    "ScreamTail": "Scream Tail",
    "SandyShock": "Sandy Shock",
    "Flut. Mane": "Flutter Mane",
    "RagingBolt": "Raging Bolt",
    "GouginFire": "Gouging Fire",
    "Roar. Moon": "Roaring Moon",
    "WalkinWake": "Walking Wake",
}


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_space(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "")
    value = value.replace("\u2019", "'").replace("\u2018", "'")
    value = value.replace("\u201c", '"').replace("\u201d", '"')
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def strip_stars(value: str) -> str:
    value = value.replace("⭐", "")
    value = value.replace("( )", "")
    value = re.sub(r"\(\s*\)", "", value)
    return clean_space(value)


def clean_species_name(raw: str) -> str:
    value = clean_space(as_text(raw))
    value = value.replace("(⭐)", "")
    value = value.replace("⭐", "")
    value = re.sub(r"^\s*-\s*", "", value)
    value = value.replace("Mr.Mime", "Mr. Mime")
    value = value.replace("Porygon 2", "Porygon2")
    value = value.replace("Sirfetch'd", "Sirfetch'd")
    value = value.replace("Sirfetch’d", "Sirfetch'd")
    value = re.sub(r"\s+", " ", value)
    value = NAME_ALIASES.get(value, value)
    return value.strip()


def name_key(raw: str) -> str:
    value = clean_species_name(raw).lower()
    value = re.sub(r"^b\.?b\.?\s+", "bb ", value)
    value = value.replace("♀", " f").replace("♂", " m")
    value = value.replace("(f)", " f").replace("(m)", " m")
    value = value.replace("nidoran (f)", "nidoran f")
    value = value.replace("nidoran (m)", "nidoran m")
    value = value.replace("mr. mime", "mr mime")
    value = value.replace("mime jr.", "mime jr")
    value = value.replace("porygon-z", "porygon z")
    value = value.replace("porygon z", "porygon-z")
    value = value.replace("sirfetch’d", "sirfetchd")
    value = value.replace("sirfetch'd", "sirfetchd")
    value = value.replace("farfetch’d", "farfetchd")
    value = value.replace("farfetch'd", "farfetchd")
    value = value.replace("battle bond", "bb")
    value = value.replace("b.b.", "bb").replace("b.b", "bb")
    if value.startswith("bb "):
        value = f"{value[3:]} bb"
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    return value


def display_name(raw: str, is_variant: bool = False) -> str:
    name = clean_species_name(raw)
    if is_variant and "⭐" not in name:
        return f"{name} ⭐"
    return name


def split_types(value: str) -> list[str]:
    return [part.strip().title() for part in clean_space(as_text(value)).split("/") if part.strip()]


def split_abilities(value: str) -> list[str]:
    return [part.strip() for part in clean_space(as_text(value)).split("/") if part.strip()]


def ability_key(value: str) -> str:
    value = clean_space(value).lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", value)


def is_variant(raw: str) -> bool:
    text = clean_space(as_text(raw))
    return "⭐" in text and "(⭐)" not in text


def evolves_to_variant(raw: str) -> bool:
    return "(⭐)" in clean_space(as_text(raw))


def level_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.day}-{value.month}"
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return clean_space(str(value))


def encounter_rate_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if 0 <= value <= 1:
            pct = value * 100
            return f"{pct:.1f}%".replace(".0%", "%")
        return str(int(value)) if value.is_integer() else str(value)
    return clean_space(str(value))


def min_level(level: str) -> int | None:
    nums = [int(x) for x in re.findall(r"\d+", level or "")]
    return min(nums) if nums else None


def timing_for_location(location: str, method: str = "") -> str:
    text = f"{location} {method}".upper()
    if "POSTGAME" in text:
        return "Postgame"
    for needle in LATE_LOCATIONS:
        if needle in text:
            return "Late"
    for needle in MID_LOCATIONS:
        if needle in text:
            return "Mid"
    for needle in EARLY_LOCATIONS:
        if needle in text:
            return "Early"
    return "Unknown"


def timing_rank(timing: str) -> int:
    return TIMING_ORDER.get(timing, TIMING_ORDER["Unknown"])


def checkpoint_for_id(checkpoint_id: str | None) -> dict:
    return CHECKPOINT_BY_ID.get(checkpoint_id or "", CHECKPOINT_BY_ID["unknown"])


def checkpoint_for_sort(sort: int | None) -> dict:
    if sort is None:
        return CHECKPOINT_BY_ID["unknown"]
    return CHECKPOINT_BY_SORT.get(sort, CHECKPOINT_BY_ID["unknown"])


def checkpoint_for_level(level: int | None) -> dict:
    if level is None or level <= 0:
        return CHECKPOINT_BY_ID["first-stratum"]
    for checkpoint in PROGRESSION_CHECKPOINTS:
        cap = checkpoint.get("levelCap")
        if cap is not None and cap >= level:
            return checkpoint
    return CHECKPOINT_BY_ID["postgame"]


def checkpoint_for_location(location: str, area_map: dict[str, str] | None = None, default: str = "unknown") -> dict:
    text = clean_space(location).upper()
    if not text:
        return checkpoint_for_id(default)
    if "POSTGAME" in text:
        return CHECKPOINT_BY_ID["postgame"]
    after_matches = [
        ("AFTER THE FIRST STRATUM", 2),
        ("AFTER FIRST STRATUM", 2),
        ("AFTER THE SECOND STRATUM", 3),
        ("AFTER SECOND STRATUM", 3),
        ("AFTER THE THIRD STRATUM", 4),
        ("AFTER THIRD STRATUM", 4),
        ("AFTER THE FOURTH STRATUM", 5),
        ("AFTER FOURTH STRATUM", 5),
        ("AFTER THE FIFTH STRATUM", 6),
        ("AFTER FIFTH STRATUM", 6),
        ("AFTER THE SIXTH STRATUM", 7),
        ("AFTER SIXTH STRATUM", 7),
        ("AFTER THE SEVENTH STRATUM", 8),
        ("AFTER SEVENTH STRATUM", 8),
    ]
    for needle, sort in after_matches:
        if needle in text:
            return checkpoint_for_sort(sort)
    if area_map:
        for area, checkpoint_id in sorted(area_map.items(), key=lambda item: len(item[0]), reverse=True):
            if area and area in text:
                return checkpoint_for_id(checkpoint_id)
    for sort, needles in LOCATION_CHECKPOINT_RULES:
        if any(needle in text for needle in needles):
            return checkpoint_for_sort(sort)
    return checkpoint_for_id(default)


def checkpoint_payload(checkpoint: dict) -> dict:
    return {
        "id": checkpoint["id"],
        "sort": checkpoint["sort"],
        "label": checkpoint["label"],
        "shortLabel": checkpoint["shortLabel"],
        "levelCap": checkpoint.get("levelCap"),
    }


def max_checkpoint(*checkpoints: dict) -> dict:
    return max((checkpoint for checkpoint in checkpoints if checkpoint), key=lambda item: item["sort"], default=CHECKPOINT_BY_ID["unknown"])


def item_key(raw: str) -> str:
    value = clean_space(raw).lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", value)


def canonical_item_name(raw: str, item_names: set[str] | None = None) -> str:
    key = item_key(raw)
    if key in ITEM_NAME_ALIASES:
        return ITEM_NAME_ALIASES[key]
    if item_names:
        by_key = {item_key(item): item for item in item_names}
        if key in by_key:
            return by_key[key]
    return clean_space(raw)


def is_area_heading(value: str, row_values: list[str]) -> bool:
    text = clean_space(value)
    if not text:
        return False
    upper = text.upper()
    if upper in METHOD_TOKENS or upper in HEADER_TOKENS:
        return False
    if "POKÉMON" in upper or "POKEMON" in upper or "LEVEL" in upper or "ENCOUNTER" in upper:
        return False
    nonempty = [clean_space(v) for v in row_values if clean_space(v)]
    if len(nonempty) > 2:
        return False
    return upper == text and len(text) > 2


def is_method_heading(value: str) -> bool:
    text = clean_space(value).upper()
    return text in METHOD_TOKENS or text.startswith("F.O.E")


def is_encounter_name(value: str) -> bool:
    text = clean_space(value)
    if not text:
        return False
    upper = text.upper()
    if upper in METHOD_TOKENS or upper in HEADER_TOKENS:
        return False
    if "POKÉMON" in upper or "POKEMON" in upper:
        return False
    return True


def split_species_list(raw: str) -> list[str]:
    value = clean_space(as_text(raw))
    if not value:
        return []
    value = value.replace(" or ", "/").replace(" and ", "/").replace(",", "/")
    value = re.sub(r"\s*/\s*", "/", value)
    names = [clean_species_name(part) for part in value.split("/") if clean_species_name(part)]
    return names


def normalize_evolution_part(part: str) -> str:
    part = clean_space(part)
    level_match = re.search(r"LV\.?\s*(\d+)", part, flags=re.I)
    if level_match:
        return f"Lv. {level_match.group(1)}"
    return part


def parse_evolution_method(raw: str) -> dict:
    raw = clean_space(as_text(raw))
    if not raw or raw in {"/", "➜"}:
        return {
            "raw": raw,
            "kind": "none",
            "label": "Final stage",
            "items": [],
            "levels": [],
            "isItemBased": False,
            "isStoneBased": False,
        }
    if raw in SPECIAL_EVOLUTION_GATES:
        gate = SPECIAL_EVOLUTION_GATES[raw]
        checkpoint = checkpoint_for_id(gate["checkpointId"])
        return {
            "raw": raw,
            "kind": "special",
            "label": f"Evolves at {raw}",
            "items": [],
            "levels": [],
            "isItemBased": False,
            "isStoneBased": False,
            "specialGate": {
                "label": raw,
                "source": gate["source"],
                "note": gate["note"],
                "checkpoint": checkpoint_payload(checkpoint),
            },
        }

    parts = [normalize_evolution_part(part) for part in re.split(r"\s*/\s*", raw) if clean_space(part)]
    levels = [part for part in parts if re.match(r"Lv\.\s*\d+$", part, flags=re.I)]
    happiness = [part for part in parts if "happiness" in part.lower() or "friendship" in part.lower()]
    items = [part for part in parts if part not in levels and part not in happiness]
    is_stone = any(any(term.lower() == item.lower() for term in STONE_EVOLUTION_TERMS) or "stone" in item.lower() for item in items)

    if items and (levels or happiness):
        kind = "mixed"
    elif items:
        kind = "item"
    elif levels:
        kind = "level"
    elif happiness:
        kind = "friendship"
    else:
        kind = "special"

    if kind == "level":
        label = f"Evolves at {levels[0]}" if len(levels) == 1 else f"Evolves at {' / '.join(levels)}"
    elif kind == "friendship":
        label = "Evolves with high friendship"
    elif kind == "mixed":
        label = f"Evolves by {' / '.join([*levels, *happiness, *items])}"
    elif items:
        label = f"Evolves with {' / '.join(items)}"
    else:
        label = f"Evolves by {raw}"

    return {
        "raw": raw,
        "kind": kind,
        "label": label,
        "items": items,
        "levels": levels,
        "isItemBased": bool(items),
        "isStoneBased": is_stone,
    }


def evolution_method_summary(method: dict) -> str:
    parts = [*method.get("levels", []), *method.get("items", [])]
    return " / ".join(parts) or method.get("raw") or method.get("label", "")


def incoming_evolution_payload(pokemon: dict, source_key: str, raw_method: str) -> dict:
    source = pokemon[source_key]
    method = parse_evolution_method(raw_method)
    return {
        **method,
        "fromId": source_key,
        "fromName": source["displayName"],
        "label": f"From {source['displayName']}: {evolution_method_summary(method)}",
    }


def ensure_pokemon(pokemon: dict, raw_name: str, source: str = "") -> dict:
    key = name_key(raw_name)
    if not key:
        return {}
    if key not in pokemon:
        pokemon[key] = {
            "id": key,
            "name": clean_species_name(raw_name),
            "displayName": clean_species_name(raw_name),
            "isVariant": is_variant(raw_name),
            "evolvesToVariant": evolves_to_variant(raw_name),
            "types": [],
            "abilities": [],
            "evolution": "",
            "evolutionMethod": None,
            "incomingEvolution": None,
            "learnset": [],
            "stats": None,
            "vanillaStats": None,
            "statDelta": None,
            "baseTotal": None,
            "vanillaTotal": None,
            "sprite": None,
            "gameDexId": None,
            "family": [],
            "directEncounters": [],
            "familyEncounters": [],
            "guideEntryIds": [],
            "roles": [],
            "recommendedSets": [],
            "availability": {
                "phase": "Unknown",
                "sort": timing_rank("Unknown"),
                "source": "No encounter source parsed yet",
                "details": "",
                "via": "",
            },
            "timeline": None,
            "sourceSheets": [],
        }
    mon = pokemon[key]
    if is_variant(raw_name):
        mon["isVariant"] = True
        mon["displayName"] = display_name(mon["name"], True)
    if evolves_to_variant(raw_name):
        mon["evolvesToVariant"] = True
    if source and source not in mon["sourceSheets"]:
        mon["sourceSheets"].append(source)
    return mon


def parse_pokedex(wb, pokemon: dict) -> None:
    ws = wb["Pokédex"]
    for row in range(3, ws.max_row + 1):
        dex_id = as_text(ws.cell(row, 2).value)
        name = as_text(ws.cell(row, 3).value)
        if not dex_id or not name:
            continue
        mon = ensure_pokemon(pokemon, name, "Pokédex")
        mon["gameDexId"] = dex_id
        if ws.cell(row, 1).value:
            mon["availability"]["source"] = clean_space(as_text(ws.cell(row, 1).value)).title()


def parse_stats_and_learnsets(wb, pokemon: dict) -> list[list[str]]:
    family_groups: list[list[str]] = []
    for sheet_name in STATS_SHEETS:
        ws = wb[sheet_name]
        headers_by_col: dict[int, list[int]] = defaultdict(list)
        for col in POKEMON_START_COLS:
            for row in range(1, ws.max_row):
                if ws.cell(row, col).value and clean_space(as_text(ws.cell(row + 1, col).value)).lower() == "type:":
                    headers_by_col[col].append(row)

        row_to_family: dict[int, list[str]] = defaultdict(list)
        for col, rows in headers_by_col.items():
            for idx, row in enumerate(rows):
                raw_name = as_text(ws.cell(row, col).value)
                mon = ensure_pokemon(pokemon, raw_name, sheet_name)
                if not mon:
                    continue
                mon["types"] = split_types(ws.cell(row + 1, col + 1).value)
                mon["abilities"] = split_abilities(ws.cell(row + 2, col + 1).value)
                mon["evolution"] = clean_space(as_text(ws.cell(row + 3, col + 1).value))
                mon["evolutionMethod"] = parse_evolution_method(mon["evolution"])

                next_row = rows[idx + 1] if idx + 1 < len(rows) else ws.max_row + 1
                moves = []
                seen_moves = set()
                for move_row in range(row + 5, next_row):
                    level = clean_space(as_text(ws.cell(move_row, col).value))
                    move = clean_space(as_text(ws.cell(move_row, col + 1).value))
                    if not level or not move or move == "/":
                        continue
                    if not re.search(r"\d|LV|Evo|Tutor|TM", level, flags=re.I):
                        continue
                    sig = (level, move)
                    if sig in seen_moves:
                        continue
                    seen_moves.add(sig)
                    moves.append({"level": level.replace("LV.", "Lv.").replace("LV", "Lv."), "move": move})
                mon["learnset"] = moves
                row_to_family[row].append(mon["id"])

        if sheet_name.startswith("#"):
            for members in row_to_family.values():
                unique_members = []
                for key in members:
                    if key not in unique_members:
                        unique_members.append(key)
                if len(unique_members) > 1:
                    family_groups.append(unique_members)

        for row in range(1, ws.max_row - 3):
            raw_name = as_text(ws.cell(row, 10).value)
            if not raw_name:
                continue
            if clean_space(as_text(ws.cell(row + 1, 10).value)).upper() != "HP":
                continue
            if clean_space(as_text(ws.cell(row + 2, 17).value)).lower() != "odyssey":
                continue
            mon = ensure_pokemon(pokemon, raw_name, sheet_name)
            stats = {}
            vanilla = {}
            for index, stat in enumerate(STAT_NAMES, start=10):
                if isinstance(ws.cell(row + 2, index).value, (int, float)):
                    stats[stat] = int(ws.cell(row + 2, index).value)
                if isinstance(ws.cell(row + 3, index).value, (int, float)):
                    vanilla[stat] = int(ws.cell(row + 3, index).value)
            total = ws.cell(row + 2, 16).value
            vanilla_total = ws.cell(row + 3, 16).value
            mon["stats"] = stats or None
            mon["vanillaStats"] = vanilla or None
            mon["baseTotal"] = int(total) if isinstance(total, (int, float)) else None
            mon["vanillaTotal"] = int(vanilla_total) if isinstance(vanilla_total, (int, float)) else None
            if stats and vanilla:
                mon["statDelta"] = {stat: stats.get(stat, 0) - vanilla.get(stat, 0) for stat in STAT_NAMES}
                if mon["baseTotal"] is not None and mon["vanillaTotal"] is not None:
                    mon["statDelta"]["total"] = mon["baseTotal"] - mon["vanillaTotal"]
    return family_groups


def image_anchor_cell(image) -> tuple[int, int] | None:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if not marker:
        return None
    return marker.row + 1, marker.col + 1


def write_clean_png(image_data: bytes, path: Path) -> None:
    with Image.open(BytesIO(image_data)) as image:
        clean_image = image.convert("RGBA") if image.mode not in {"RGB", "RGBA"} else image.copy()
    clean_image.save(path, format="PNG", optimize=True)


def parse_etrian_variant_sprites(wb, pokemon: dict) -> tuple[int, list[str]]:
    if "Etrian Variants" not in wb.sheetnames:
        return 0, []

    SPRITE_ASSET_DIR.mkdir(parents=True, exist_ok=True)
    for old_asset in SPRITE_ASSET_DIR.glob("*.png"):
        old_asset.unlink()

    ws = wb["Etrian Variants"]
    extracted = 0
    unmatched: list[str] = []
    seen: set[str] = set()

    for image in getattr(ws, "_images", []):
        anchor = image_anchor_cell(image)
        if not anchor:
            continue
        row, col = anchor
        if clean_space(as_text(ws.cell(row, 1).value)).upper() != "NORMAL":
            continue
        raw_name = clean_space(as_text(ws.cell(row - 1, col).value))
        if not raw_name:
            continue
        key = resolve_pokemon_key(name_key(raw_name), pokemon)
        if not key:
            unmatched.append(raw_name)
            continue
        if key in seen:
            continue
        seen.add(key)
        asset_path = SPRITE_ASSET_DIR / f"{key}.png"
        write_clean_png(image._data(), asset_path)
        pokemon[key]["sprite"] = {
            "source": "Etrian Variants",
            "kind": "normal",
            "path": f"{SPRITE_ASSET_PREFIX}/{key}.png",
            "sourceName": raw_name,
        }
        extracted += 1

    return extracted, unmatched


def parse_ability_definitions(wb) -> list[dict]:
    ws = wb["New Moves & Abilities"]
    ability_columns = [2, 5, 8, 11, 14, 17]
    definitions = {}
    current_section = ""

    for row in range(1, ws.max_row):
        first = clean_space(as_text(ws.cell(row, 1).value)).upper()
        if first in {"NEW ABILITIES", "BUFFED ABILITIES"}:
            current_section = first.title()
            continue
        if first == "BUFFED/REWORKED MOVES":
            current_section = ""
            continue
        if not current_section:
            continue

        for col in ability_columns:
            raw_name = clean_space(as_text(ws.cell(row, col).value))
            effect = clean_space(as_text(ws.cell(row + 1, col).value))
            if not raw_name or not effect:
                continue
            if raw_name.upper() in {"TYPE", "CATEGORY", "POWER", "ACCURACY", "PP", "EFFECT"}:
                continue
            if len(raw_name) < 3:
                continue
            key = ability_key(raw_name)
            definitions[key] = {
                "id": key,
                "name": raw_name.title(),
                "effect": effect,
                "section": current_section,
                "source": "New Moves & Abilities",
            }

    return sorted(definitions.values(), key=lambda item: item["name"])


def ability_names_from_pokemon(pokemon: dict) -> set[str]:
    names = set()
    for mon in pokemon.values():
        for ability in mon.get("abilities", []):
            if ability:
                names.add(ability)
    return names


def http_json(url: str) -> dict:
    request = urllib.request.Request(url, headers=HTTP_HEADERS)
    with urllib.request.urlopen(request, timeout=20) as response:
        return json.load(response)


def load_standard_ability_cache() -> dict[str, dict]:
    if not STANDARD_ABILITY_CACHE.exists():
        return {}
    try:
        raw = json.loads(STANDARD_ABILITY_CACHE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return {ability_key(entry.get("name", entry.get("id", ""))): entry for entry in raw if entry.get("effect")}


def save_standard_ability_cache(cache: dict[str, dict]) -> None:
    STANDARD_ABILITY_CACHE.parent.mkdir(parents=True, exist_ok=True)
    entries = sorted(cache.values(), key=lambda item: item["name"])
    STANDARD_ABILITY_CACHE.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")


def english_short_effect(data: dict) -> str:
    for entry in data.get("effect_entries", []):
        if entry.get("language", {}).get("name") == "en" and entry.get("short_effect"):
            effect = clean_space(entry["short_effect"])
            chance = data.get("effect_chance")
            if chance is not None:
                effect = effect.replace("$effect_chance%", f"{chance}%")
                effect = effect.replace("$effect_chance", str(chance))
            return effect
    for entry in data.get("flavor_text_entries", []):
        if entry.get("language", {}).get("name") == "en" and entry.get("flavor_text"):
            return clean_space(entry["flavor_text"])
    return ""


def fetch_standard_ability_definitions(ability_names: set[str], odyssey_keys: set[str]) -> list[dict]:
    needed = {ability_key(name): name for name in ability_names if ability_key(name) not in odyssey_keys}
    cache = load_standard_ability_cache()
    missing = sorted(key for key in needed if key not in cache)

    if missing:
        try:
            index = http_json(POKEAPI_ABILITY_INDEX)
            urls = {ability_key(item["name"]): item["url"] for item in index.get("results", [])}
            targets = []
            for key in missing:
                lookup_key = STANDARD_ABILITY_ALIASES.get(key, key)
                if lookup_key in urls:
                    targets.append((key, needed[key], urls[lookup_key]))
            if targets:
                with ThreadPoolExecutor(max_workers=8) as executor:
                    futures = {
                        executor.submit(http_json, url): (key, name)
                        for key, name, url in targets
                    }
                    for future in as_completed(futures):
                        key, name = futures[future]
                        try:
                            data = future.result()
                        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError):
                            continue
                        effect = english_short_effect(data)
                        if not effect:
                            continue
                        cache[key] = {
                            "id": key,
                            "name": name,
                            "effect": effect,
                            "section": "Standard Ability",
                            "source": "PokeAPI",
                        }
                save_standard_ability_cache(cache)
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
            print(f"Warning: could not refresh standard ability definitions: {exc}")

    return sorted(
        [entry for key, entry in cache.items() if key in needed],
        key=lambda item: item["name"],
    )


def merge_ability_definitions(odyssey_definitions: list[dict], standard_definitions: list[dict]) -> list[dict]:
    merged = {entry["id"]: entry for entry in standard_definitions}
    for entry in odyssey_definitions:
        merged[entry["id"]] = entry
    return sorted(merged.values(), key=lambda item: item["name"])


def parse_level_caps(wb) -> list[dict]:
    ws = wb["Level Caps"]
    caps = []
    for row in range(1, ws.max_row + 1):
        stratum = clean_space(as_text(ws.cell(row, 1).value))
        cap = clean_space(as_text(ws.cell(row, 3).value))
        if not stratum or not cap or not cap.upper().startswith("LV"):
            continue
        checkpoint = checkpoint_for_location(stratum)
        caps.append({
            "stratum": stratum,
            "cap": cap.replace("LV.", "Lv.").replace("LV", "Lv."),
            "phase": timing_for_location(stratum),
            "checkpoint": checkpoint_payload(checkpoint),
        })
    return caps


def make_encounter_record(source: str, area: str, method: str, raw_name: str, level, rate, checkpoint_id: str | None = None, area_map: dict[str, str] | None = None) -> list[dict]:
    records = []
    level_text = level_to_text(level)
    rate_text = encounter_rate_to_text(rate)
    checkpoint = checkpoint_for_id(checkpoint_id) if checkpoint_id else checkpoint_for_location(area, area_map)
    for name in split_species_list(raw_name):
        if not name:
            continue
        records.append({
            "pokemon": name,
            "pokemonKey": name_key(name),
            "isVariant": is_variant(raw_name),
            "evolvesToVariant": evolves_to_variant(raw_name),
            "level": level_text,
            "minLevel": min_level(level_text),
            "rate": rate_text,
            "area": clean_space(area),
            "method": clean_space(method) or "Encounter",
            "phase": timing_for_location(area, method),
            "checkpoint": checkpoint["id"],
            "checkpointSort": checkpoint["sort"],
            "checkpointLabel": checkpoint["label"],
            "levelCap": checkpoint.get("levelCap"),
            "source": source,
        })
    return records


def build_area_checkpoint_map(wb) -> dict[str, str]:
    ws = wb["Pokémon"]
    area_map: dict[str, str] = {}
    for row in range(1, ws.max_row + 1):
        row_values = [as_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        first = clean_space(row_values[0] if row_values else "")
        if not is_area_heading(first, row_values):
            continue
        checkpoint = checkpoint_for_location(first)
        area_map[first.upper()] = checkpoint["id"]
    return area_map


def parse_wild_encounters(wb, area_map: dict[str, str] | None = None) -> list[dict]:
    ws = wb["Pokémon"]
    encounters = []
    current_area = ""
    current_checkpoint_id = "first-stratum"
    current_methods: dict[int, str] = defaultdict(str)
    for row in range(1, ws.max_row + 1):
        row_values = [as_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        first = clean_space(row_values[0] if row_values else "")
        if is_area_heading(first, row_values):
            current_area = first
            current_checkpoint_id = checkpoint_for_location(first, area_map, current_checkpoint_id)["id"]
            current_methods = defaultdict(str)
            continue

        for col in ENCOUNTER_START_COLS:
            if col > ws.max_column:
                continue
            value = clean_space(as_text(ws.cell(row, col).value))
            if is_method_heading(value):
                current_methods[col] = value.title() if value.isupper() else value

        for col in ENCOUNTER_START_COLS:
            if col + 2 > ws.max_column:
                continue
            raw_name = as_text(ws.cell(row, col).value)
            if not is_encounter_name(raw_name):
                continue
            level = ws.cell(row, col + 1).value
            rate = ws.cell(row, col + 2).value
            if level is None and rate is None:
                continue
            method = current_methods[col] or "Encounter"
            encounters.extend(make_encounter_record("Wild encounters", current_area, method, raw_name, level, rate, current_checkpoint_id, area_map))

    post = wb["Pokémon (Postgame)"]
    for row in range(1, post.max_row + 1):
        for col in [2, 6, 10]:
            if col + 2 > post.max_column:
                continue
            raw_name = as_text(post.cell(row, col).value)
            if not is_encounter_name(raw_name):
                continue
            level = post.cell(row, col + 1).value
            rate = post.cell(row, col + 2).value
            if level is None and rate is None:
                continue
            area = ""
            for scan in range(row, 0, -1):
                possible = as_text(post.cell(scan, col).value)
                if possible and possible != raw_name and not is_method_heading(possible) and "POKÉMON" not in possible.upper():
                    area = possible
                    break
            records = make_encounter_record("Postgame encounters", area or "Postgame", "Event (Postgame)", raw_name, level, rate, "postgame", area_map)
            for record in records:
                record["phase"] = "Postgame"
            encounters.extend(records)
    return encounters


def parse_wonder_trade(wb) -> list[dict]:
    ws = wb["Wonder Trade"]
    encounters = []
    groups = [(2, "Kanto Wonder Trade"), (5, "Johto Wonder Trade"), (8, "Hoenn Wonder Trade")]
    for row in range(5, ws.max_row + 1):
        for col, area in groups:
            raw_name = as_text(ws.cell(row, col).value)
            if not raw_name:
                continue
            records = make_encounter_record("Wonder Trade", area, "Wonder Trade", raw_name, ws.cell(row, col + 1).value, "Trade", "first-stratum")
            for record in records:
                record["phase"] = "Start"
                record["startAvailable"] = True
            encounters.extend(records)
    return encounters


def parse_naval_explorations(wb) -> list[dict]:
    ws = wb["Naval Explorations"]
    encounters = []
    current_phase = "Naval Exploration"
    current_methods: dict[int, str] = defaultdict(str)
    for row in range(1, ws.max_row + 1):
        first = clean_space(as_text(ws.cell(row, 1).value))
        if first and ("Encounter" not in first) and (first == "Normal Encounters" or first.startswith("After ")):
            current_phase = first
            current_methods = defaultdict(str)
            continue
        for col in [1, 4, 7, 11]:
            if col > ws.max_column:
                continue
            value = clean_space(as_text(ws.cell(row, col).value))
            if value.startswith("ENCOUNTER") or value.startswith("F.O.E"):
                current_methods[col] = value
        for col in [1, 4, 7, 11]:
            if col + 1 > ws.max_column:
                continue
            raw_name = as_text(ws.cell(row, col).value)
            if not is_encounter_name(raw_name):
                continue
            level = ws.cell(row, col + 1).value
            if level is None:
                continue
            rate = ws.cell(row, col + 2).value if col + 2 <= ws.max_column else ""
            checkpoint = checkpoint_for_location(current_phase, default="first-stratum" if current_phase == "Normal Encounters" else "unknown")
            records = make_encounter_record("Naval Explorations", current_phase, current_methods[col] or "Naval Encounter", raw_name, level, rate, checkpoint["id"])
            for record in records:
                if current_phase == "Normal Encounters":
                    record["phase"] = "Early"
                else:
                    record["phase"] = timing_for_location(current_phase)
            encounters.extend(records)
    return encounters


def parse_tms_and_tutors(wb, area_map: dict[str, str] | None = None) -> tuple[list[dict], list[dict]]:
    tms = []
    ws = wb["TM Location"]
    for row in range(4, ws.max_row + 1):
        number = clean_space(as_text(ws.cell(row, 2).value))
        move = clean_space(as_text(ws.cell(row, 3).value))
        location = clean_space(as_text(ws.cell(row, 4).value))
        if number and move:
            checkpoint = checkpoint_for_location(location, area_map)
            tms.append({"number": number, "move": move, "location": location, "phase": timing_for_location(location), "checkpoint": checkpoint_payload(checkpoint)})

    tutors = []
    ws = wb["Move Tutors"]
    for row in range(4, ws.max_row + 1):
        move = clean_space(as_text(ws.cell(row, 2).value))
        location = clean_space(as_text(ws.cell(row, 3).value))
        if move:
            checkpoint = checkpoint_for_location(location, area_map)
            tutors.append({"move": move, "location": location, "phase": timing_for_location(location), "checkpoint": checkpoint_payload(checkpoint)})
    return tms, tutors


def parse_sidequests(wb, area_map: dict[str, str] | None = None) -> list[dict]:
    ws = wb["Sidequests"]
    quests = []
    for row in range(2, ws.max_row + 1):
        number = clean_space(as_text(ws.cell(row, 1).value))
        name = clean_space(as_text(ws.cell(row, 2).value))
        location = clean_space(as_text(ws.cell(row, 4).value))
        description = clean_space(as_text(ws.cell(row, 6).value))
        reward = clean_space(as_text(ws.cell(row, 8).value))
        if not number or not name:
            continue
        quests.append({
            "number": number,
            "name": name,
            "location": location,
            "description": description,
            "reward": reward,
            "phase": timing_for_location(location),
            "checkpoint": checkpoint_payload(checkpoint_for_location(location, area_map)),
        })
    return quests


def add_item_source(sources: dict[str, list[dict]], item_names: set[str], raw_item: str, source: str, location: str, method: str, checkpoint: dict, gate_level: int | None = None, note: str = "") -> None:
    item = canonical_item_name(raw_item, item_names)
    if item_key(item) not in {item_key(name) for name in item_names}:
        return
    level_checkpoint = checkpoint_for_level(gate_level)
    final_checkpoint = max_checkpoint(checkpoint, level_checkpoint)
    sources[item].append({
        "item": item,
        "source": source,
        "location": clean_space(location),
        "method": clean_space(method),
        "checkpoint": checkpoint_payload(final_checkpoint),
        "areaCheckpoint": checkpoint_payload(checkpoint),
        "gateLevel": gate_level,
        "note": note,
    })


def foe_levels_by_area(encounters: list[dict]) -> dict[str, int]:
    levels: dict[str, int] = {}
    for record in encounters:
        if "F.O.E" not in record.get("method", "").upper():
            continue
        level = record.get("minLevel")
        if level is None:
            continue
        area = clean_space(record.get("area", "")).upper()
        if not area:
            continue
        levels[area] = min(levels.get(area, level), level)
    return levels


def foe_level_for_area(foe_levels: dict[str, int], area: str) -> int | None:
    key = clean_space(area).upper()
    if key in foe_levels:
        return foe_levels[key]
    matches = [level for foe_area, level in foe_levels.items() if key and (key in foe_area or foe_area in key)]
    return min(matches) if matches else None


def parse_field_item_sources(wb, item_names: set[str], area_map: dict[str, str], encounters: list[dict], sources: dict[str, list[dict]]) -> None:
    ws = wb["Items"]
    foe_levels = foe_levels_by_area(encounters)
    current_areas: dict[int, str] = {}
    for row in range(1, ws.max_row + 1):
        for start_col in [1, 6]:
            area = clean_space(as_text(ws.cell(row, start_col).value))
            item = clean_space(as_text(ws.cell(row, start_col + 1).value))
            method = clean_space(as_text(ws.cell(row, start_col + 2).value))
            if area and not item and not method:
                current_areas[start_col] = area
                continue
            if not item:
                continue
            current_area = current_areas.get(start_col, "")
            checkpoint = checkpoint_for_location(current_area, area_map)
            gate_level = None
            note = ""
            if "F.O.E" in method.upper():
                gate_level = foe_level_for_area(foe_levels, current_area)
                note = "Guarded by an F.O.E.; checkpoint uses the parsed F.O.E. level when available."
                if gate_level is None:
                    checkpoint = CHECKPOINT_BY_ID["unknown"]
                    note = "Guarded by an F.O.E.; no guard level was parsed, so this source is not used for early timeline claims."
            add_item_source(sources, item_names, item, "Field item", current_area, method, checkpoint, gate_level, note)


def parse_gathering_mining_item_sources(wb, item_names: set[str], sources: dict[str, list[dict]]) -> None:
    ws = wb["GatheringMining"]
    current_checkpoint = CHECKPOINT_BY_ID["unknown"]
    current_methods: dict[int, str] = {}
    for row in range(1, ws.max_row + 1):
        first = clean_space(as_text(ws.cell(row, 1).value))
        if first and "STRATUM" in first.upper():
            current_checkpoint = checkpoint_for_location(first)
            current_methods = {}
            continue
        for col in [1, 4, 7]:
            header = clean_space(as_text(ws.cell(row, col).value))
            if header and any(token in header.upper() for token in ["GATHERING", "MINING"]):
                current_methods[col] = header
                continue
            item = header
            if item:
                add_item_source(sources, item_names, item.replace(" ↑↑", ""), "Gathering/Mining", current_checkpoint["label"], current_methods.get(col, "Gathering/Mining"), current_checkpoint)


def parse_napier_item_sources(wb, item_names: set[str], area_map: dict[str, str], sources: dict[str, list[dict]]) -> None:
    ws = wb["Items (Shop)"]
    napier_checkpoint = checkpoint_for_location("Aqua Resort", area_map)
    tier_by_col: dict[int, int] = {}
    in_napier = False
    for row in range(1, ws.max_row + 1):
        row_text = " ".join(clean_space(as_text(ws.cell(row, col).value)) for col in range(1, ws.max_column + 1))
        if "NAPIER'S SHOP" in row_text.upper():
            in_napier = True
            continue
        if in_napier and "POWER ITEM SHOP" in row_text.upper():
            break
        if not in_napier:
            continue
        for col in range(1, ws.max_column + 1):
            value = clean_space(as_text(ws.cell(row, col).value))
            match = re.search(r"TREASURES\s+OBTAINED:\s*(\d+)", value, flags=re.I)
            if match:
                tier_by_col[col] = int(match.group(1))
                continue
            if not value:
                continue
            tier = None
            for header_col, header_tier in sorted(tier_by_col.items()):
                if col >= header_col and col < header_col + 3:
                    tier = header_tier
                    break
            if tier is None:
                continue
            gate_level = NAPIER_TREASURE_TIER_LEVELS.get(tier)
            note = f"Napier shop tier {tier}; tier gates are conservative assumptions from documented sea boss/F.O.E. levels."
            add_item_source(sources, item_names, value, "Napier's Shop", "Aqua Resort", f"Treasures obtained: {tier}", napier_checkpoint, gate_level, note)


def parse_sidequest_item_sources(sidequests: list[dict], item_names: set[str], sources: dict[str, list[dict]]) -> None:
    for quest in sidequests:
        reward = quest.get("reward", "")
        for item in item_names:
            if item_key(item) and item_key(item) in item_key(reward):
                checkpoint = checkpoint_for_id(quest.get("checkpoint", {}).get("id"))
                add_item_source(sources, item_names, item, "Sidequest reward", quest.get("location", ""), quest.get("name", ""), checkpoint, note=f"Reward from sidequest {quest.get('number', '')}: {quest.get('name', '')}")


def evolution_item_names_from_pokemon(pokemon: dict) -> set[str]:
    names = set()
    for mon in pokemon.values():
        for method in [mon.get("evolutionMethod"), mon.get("incomingEvolution")]:
            if not method:
                continue
            for item in method.get("items", []):
                if item:
                    names.add(canonical_item_name(item))
    return names


def parse_evolution_item_sources(wild_wb, sidequests: list[dict], item_names: set[str], area_map: dict[str, str], encounters: list[dict]) -> list[dict]:
    sources: dict[str, list[dict]] = defaultdict(list)
    for stone in LOTTERY_STONES & item_names:
        add_item_source(
            sources,
            item_names,
            stone,
            "Talrega Lottery",
            "Talrega",
            "Lottery prize",
            CHECKPOINT_BY_ID["first-stratum"],
            note="Manual rule: Fire, Water, Thunder, and Sun Stones are available from the Talrega Lottery before the First Stratum cap.",
        )
    parse_field_item_sources(wild_wb, item_names, area_map, encounters, sources)
    parse_gathering_mining_item_sources(wild_wb, item_names, sources)
    parse_napier_item_sources(wild_wb, item_names, area_map, sources)
    parse_sidequest_item_sources(sidequests, item_names, sources)

    rows = []
    for item, item_sources in sources.items():
        item_sources.sort(key=lambda source: (source["checkpoint"]["sort"], source.get("gateLevel") or 0, source["source"], source["location"]))
        rows.extend(item_sources)
    rows.sort(key=lambda source: (source["item"], source["checkpoint"]["sort"], source.get("gateLevel") or 0))
    return rows


def normalize_role(raw_role: str) -> str:
    role = clean_space(raw_role)
    role = re.sub(r"^\d+(?:\.\d)?\s*-\s*", "", role)
    role = role.title()
    role = role.replace("Aether Types", "Aether")
    role = role.replace("Weather-Related", "Weather-related")
    role = role.replace("Follow Me/Rage Powder Users", "Redirection User")
    role = role.replace("Fake Out Users", "Fake Out User")
    role = role.replace("Intimidate Users", "Intimidate User")
    role = role.replace("Levitate Users", "Levitate User")
    role = role.replace("Sun Setters", "Sun Setter")
    role = role.replace("Rain Setters", "Rain Setter")
    role = role.replace("Sandstorm Setters", "Sandstorm Setter")
    role = role.replace("Hail Setters", "Hail Setter")
    return role


def clean_pdf_text(text: str) -> str:
    text = text.replace("\n", " ")
    text = text.replace("|", " ")
    text = clean_space(text)
    text = text.replace("Talrega ’s", "Talrega's")
    text = text.replace("Saya’s", "Saya's")
    return text


def best_heading(raw: str) -> str:
    raw = clean_space(raw)
    matches = re.findall(r"(?:\d+\.\d|\d+\.0)\s*-\s*([^●]+?)(?=(?:\d+\.\d|\d+\.0)\s*-|$)", raw)
    return clean_space(matches[-1] if matches else raw)


def parse_bullet_guide(pdf_path: Path) -> list[dict]:
    reader = PdfReader(str(pdf_path))
    entries = []
    entry_id = 1
    for page_index in range(5, 13):
        if page_index >= len(reader.pages):
            continue
        text = clean_pdf_text(reader.pages[page_index].extract_text() or "")
        heading_matches = list(re.finditer(r"((?:\d+\.\d|\d+\.0)\s*-\s*.*?)(?=●)", text))
        if not heading_matches:
            continue
        for idx, match in enumerate(heading_matches):
            role = normalize_role(best_heading(match.group(1)))
            segment_start = match.end()
            segment_end = heading_matches[idx + 1].start() if idx + 1 < len(heading_matches) else len(text)
            segment = text[segment_start:segment_end]
            for bullet in re.finditer(r"●\s*(.*?)(?=●|$)", segment):
                body = clean_space(bullet.group(1))
                parsed = re.match(
                    r"(.+?)\s+\((Early|Mid|Late)\s+Game\)\s*(?:-\s*([^○]+?))?\s*○\s*How\s+to\s+obtain\s*:\s*(.*)",
                    body,
                    flags=re.I,
                )
                if not parsed:
                    continue
                names_text, timing, move_note, obtain = parsed.groups()
                obtain = re.sub(r"\s+\d+$", "", clean_space(obtain))
                names = split_species_list(names_text)
                entries.append({
                    "id": f"guide-{entry_id}",
                    "role": role,
                    "roleGroup": role,
                    "pokemonNames": names,
                    "pokemonKeys": [name_key(name) for name in names],
                    "availability": timing.title(),
                    "howToObtain": obtain,
                    "moveNote": clean_space(move_note or ""),
                    "item": "",
                    "ability": "",
                    "moves": [],
                    "statNote": "",
                    "source": "Teambuilding guide",
                    "page": page_index + 1,
                })
                entry_id += 1
    return entries


def parse_table_cell_name(cell: str) -> tuple[str, str]:
    lines = [clean_space(line) for line in (cell or "").split("\n") if clean_space(line)]
    if not lines:
        return "", ""
    if re.match(r"^(HP|Atk|Def|Sp\. Atk|Sp\. Def|Speed)\s*:", lines[0], flags=re.I):
        return "", " ".join(lines)
    return clean_species_name(lines[0]), " ".join(lines[1:])


def parse_table_guide(pdf_path: Path) -> list[dict]:
    if pdfplumber is None:
        return []
    role_pages = {
        14: "Great Damage Dealer",
        15: "Great Damage Dealer",
        16: "Best Support",
        17: "Best Support",
        18: "Best Tank",
        19: "Best Tank",
    }
    entries = []
    entry_id = 1000
    last_entry = None
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_number, role in role_pages.items():
            tables = pdf.pages[page_number - 1].extract_tables()
            if not tables:
                continue
            for row in tables[0]:
                if not row or row[0] == "Pokémon":
                    continue
                name, stat_note = parse_table_cell_name(row[0])
                if not name and last_entry is not None:
                    if stat_note:
                        last_entry["statNote"] = clean_space(f"{last_entry['statNote']} {stat_note}")
                    if row[1]:
                        last_entry["item"] = clean_space(f"{last_entry['item']} {row[1]}")
                    if row[2]:
                        last_entry["ability"] = clean_space(f"{last_entry['ability']} {row[2]}")
                    if row[3]:
                        last_entry["moves"].extend([clean_space(x) for x in row[3].replace("\n", " ").split(",") if clean_space(x)])
                    continue
                if not name:
                    continue
                names = [name]
                entry = {
                    "id": f"guide-{entry_id}",
                    "role": role,
                    "roleGroup": role,
                    "pokemonNames": names,
                    "pokemonKeys": [name_key(name) for name in names],
                    "availability": "",
                    "howToObtain": "",
                    "moveNote": "",
                    "item": clean_space((row[1] or "").replace("\n", " ")),
                    "ability": clean_space((row[2] or "").replace("\n", " ")),
                    "moves": [clean_space(x) for x in (row[3] or "").replace("\n", " ").split(",") if clean_space(x)],
                    "statNote": stat_note,
                    "source": "Teambuilding guide table",
                    "page": page_number,
                }
                entries.append(entry)
                last_entry = entry
                entry_id += 1
    return entries


def parse_teambuilding_guide(pdf_path: Path) -> list[dict]:
    entries = parse_bullet_guide(pdf_path)
    entries.extend(parse_table_guide(pdf_path))
    return entries


class DisjointSet:
    def __init__(self):
        self.parent = {}

    def find(self, item):
        self.parent.setdefault(item, item)
        if self.parent[item] != item:
            self.parent[item] = self.find(self.parent[item])
        return self.parent[item]

    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[rb] = ra


def apply_manual_evolution_groups(pokemon: dict, groups: list[list[str]]) -> None:
    manual_groups = [
        ["oddish", "gloom", "vileplume", "bellossom"],
        ["poliwag", "poliwhirl", "poliwrath", "politoed"],
        ["slowpoke", "slowbro", "slowking"],
        ["onix", "steelix"],
        ["scyther", "kleavor", "scizor"],
        ["seadra", "kingdra"],
        ["porygon", "porygon2", "porygon-z"],
        ["tyrogue", "hitmonlee", "hitmonchan", "hitmontop"],
        ["eevee", "vaporeon", "jolteon", "flareon", "espeon", "umbreon", "leafeon", "glaceon", "yggdreon", "goreon"],
        ["zubat", "golbat", "crobat"],
        ["happiny", "chansey", "blissey"],
        ["wurmple", "silcoon", "beautifly", "cascoon", "dustox"],
        ["farfetchd-galar", "sirfetchd"],
        ["roselia", "roserade"],
        ["aipom", "ambipom"],
        ["misdreavus", "mismagius"],
        ["murkrow", "honchkrow"],
        ["sneasel", "weavile"],
        ["magnemite", "magneton", "magnezone"],
        ["lickitung", "lickilicky"],
        ["rhyhorn", "rhydon", "rhyperior"],
        ["tangela", "tangrowth"],
        ["elekid", "electabuzz", "electivire"],
        ["magby", "magmar", "magmortar"],
        ["togepi", "togetic", "togekiss"],
        ["yanma", "yanmega"],
        ["snorunt", "glalie", "froslass"],
        ["gligar", "gliscor"],
        ["ralts", "kirlia", "gardevoir", "gallade"],
        ["nosepass", "probopass"],
        ["duskull", "dusclops", "dusknoir"],
        ["swinub", "piloswine", "mamoswine"],
        ["girafarig", "farigiraf"],
        ["corsola", "reefsola"],
        ["pikachu", "raichu", "gorochu"],
        ["rattata", "ratreecate"],
        ["plusle", "plusle-bb"],
        ["minun", "minun-bb"],
    ]
    existing = set(pokemon.keys())
    for group in manual_groups:
        keys = [key for key in group if key in existing]
        if len(keys) > 1:
            groups.append(keys)


def assign_families(pokemon: dict, groups: list[list[str]]) -> None:
    apply_manual_evolution_groups(pokemon, groups)
    order_hints = {}
    for group in groups:
        for index, key in enumerate(group):
            if key in pokemon:
                order_hints[key] = index
    dsu = DisjointSet()
    for key in pokemon:
        dsu.find(key)
    for group in groups:
        if not group:
            continue
        first = group[0]
        for key in group[1:]:
            if key in pokemon:
                dsu.union(first, key)
    families = defaultdict(list)
    for key in pokemon:
        families[dsu.find(key)].append(key)
    for members in families.values():
        sorted_members = sorted(members, key=lambda key: (order_hints.get(key, 999), int(pokemon[key]["gameDexId"] or 9999)))
        names = [pokemon[key]["displayName"] for key in sorted_members]
        for key in members:
            pokemon[key]["family"] = sorted_members
            pokemon[key]["familyNames"] = names


def assign_evolution_metadata(pokemon: dict) -> None:
    seen_families = set()
    for mon in pokemon.values():
        if not mon.get("evolutionMethod"):
            mon["evolutionMethod"] = parse_evolution_method(mon.get("evolution", ""))

    for mon in pokemon.values():
        family = tuple(mon.get("family") or [mon["id"]])
        if family in seen_families:
            continue
        seen_families.add(family)
        for index, key in enumerate(family):
            if index == 0:
                continue
            source_key = None
            for candidate_key in reversed(family[:index]):
                method = pokemon[candidate_key].get("evolutionMethod") or parse_evolution_method(pokemon[candidate_key].get("evolution", ""))
                if method.get("kind") != "none":
                    source_key = candidate_key
                    break
            if not source_key:
                continue
            source = pokemon[source_key]
            source_method = source.get("evolutionMethod") or parse_evolution_method(source.get("evolution", ""))
            pokemon[key]["incomingEvolution"] = {
                **source_method,
                "fromId": source_key,
                "fromName": source["displayName"],
                "label": f"From {source['displayName']}: {evolution_method_summary(source_method)}",
            }

    for target_key, (source_key, raw_method) in DIRECT_EVOLUTION_OVERRIDES.items():
        if target_key not in pokemon or source_key not in pokemon:
            continue
        if source_key not in pokemon[target_key].get("family", []):
            continue
        pokemon[target_key]["incomingEvolution"] = incoming_evolution_payload(pokemon, source_key, raw_method)


def sort_encounter(record: dict) -> tuple:
    return (
        record.get("checkpointSort") or timing_rank(record.get("phase", "Unknown")),
        record.get("minLevel") if record.get("minLevel") is not None else 999,
        record.get("area", ""),
        record.get("method", ""),
    )


def attach_encounters(pokemon: dict, encounters: list[dict]) -> None:
    direct = defaultdict(list)
    for record in encounters:
        record_key = resolve_pokemon_key(record["pokemonKey"], pokemon)
        if record_key:
            direct[record_key].append({**record, "pokemonKey": record_key})

    for key, records in direct.items():
        pokemon[key]["directEncounters"] = sorted(records, key=sort_encounter)

    for key, mon in pokemon.items():
        family_records = []
        for family_key in mon.get("family", [key]):
            for record in direct.get(family_key, []):
                via = pokemon.get(family_key, {}).get("displayName", record["pokemon"])
                family_records.append({**record, "via": via})
        family_records = sorted(family_records, key=sort_encounter)
        mon["familyEncounters"] = family_records[:16]
        if family_records:
            first = family_records[0]
            mon["availability"] = {
                "phase": first["phase"],
                "sort": timing_rank(first["phase"]),
                "source": first["source"],
                "details": f"{first['area']} - {first['method']}",
                "via": first.get("via", first["pokemon"]),
                "level": first.get("level", ""),
                "rate": first.get("rate", ""),
            }

    for starter in ["plusle", "minun"]:
        if starter in pokemon:
            pokemon[starter]["availability"] = {
                "phase": "Starter",
                "sort": 0,
                "source": "Story starter",
                "details": "Nyx starts the journey with Plusle and Minun.",
                "via": pokemon[starter]["displayName"],
                "level": "",
                "rate": "",
            }


def attach_guide_entries(pokemon: dict, guide_entries: list[dict]) -> None:
    alias = {}
    for key, mon in pokemon.items():
        alias[key] = key
        alias[name_key(mon["name"])] = key
        alias[name_key(mon["displayName"])] = key
        if key.endswith("-bb"):
            alias[key.replace("-bb", "")] = key

    for entry in guide_entries:
        matched = []
        for raw_key in entry.get("pokemonKeys", []):
            key = alias.get(raw_key)
            if key and key not in matched:
                matched.append(key)
        entry["pokemonKeys"] = matched or entry.get("pokemonKeys", [])
        for key in matched:
            mon = pokemon[key]
            if entry["id"] not in mon["guideEntryIds"]:
                mon["guideEntryIds"].append(entry["id"])
            if entry["role"] not in mon["roles"]:
                mon["roles"].append(entry["role"])
            if entry["source"].endswith("table"):
                mon["recommendedSets"].append({
                    "role": entry["role"],
                    "item": entry["item"],
                    "ability": entry["ability"],
                    "moves": entry["moves"],
                    "statNote": entry["statNote"],
                    "page": entry["page"],
                })

            # Guide tags are useful role hints, but concrete timeline availability
            # is derived later from encounters, level caps, and evolution gates.


def earliest_item_source(item: str, item_sources_by_key: dict[str, list[dict]]) -> dict | None:
    sources = item_sources_by_key.get(item_key(item), [])
    return sources[0] if sources else None


def method_required_level(method: dict | None) -> int | None:
    if not method:
        return None
    levels = []
    for level in method.get("levels", []):
        match = re.search(r"\d+", level)
        if match:
            levels.append(int(match.group(0)))
    return max(levels) if levels else None


def method_checkpoint(method: dict | None, item_sources_by_key: dict[str, list[dict]]) -> tuple[dict, list[dict], list[str]]:
    if not method or method.get("kind") == "none":
        return CHECKPOINT_BY_ID["first-stratum"], [], []
    checkpoints = [CHECKPOINT_BY_ID["first-stratum"]]
    item_requirements = []
    missing_items = []
    required_level = method_required_level(method)
    if required_level:
        checkpoints.append(checkpoint_for_level(required_level))
    if method.get("specialGate"):
        checkpoints.append(checkpoint_for_id(method["specialGate"]["checkpoint"]["id"]))
    for item in method.get("items", []):
        source = earliest_item_source(item, item_sources_by_key)
        if source:
            checkpoints.append(checkpoint_for_id(source["checkpoint"]["id"]))
            item_requirements.append({
                "item": source["item"],
                "source": source["source"],
                "location": source["location"],
                "method": source["method"],
                "checkpoint": source["checkpoint"],
                "gateLevel": source.get("gateLevel"),
                "note": source.get("note", ""),
            })
        else:
            missing_items.append(item)
            checkpoints.append(CHECKPOINT_BY_ID["unknown"])
    return max_checkpoint(*checkpoints), item_requirements, missing_items


def record_checkpoint(record: dict) -> dict:
    area_checkpoint = checkpoint_for_id(record.get("checkpoint"))
    level_checkpoint = checkpoint_for_level(record.get("minLevel"))
    return max_checkpoint(area_checkpoint, level_checkpoint)


def evolution_path_between(pokemon: dict, family: list[str], source_key: str, target_key: str) -> list[dict] | None:
    if source_key == target_key:
        return []
    if source_key not in family or target_key not in family:
        return None
    methods = []
    current_key = target_key
    seen = set()
    while current_key != source_key:
        if current_key in seen:
            return None
        seen.add(current_key)
        method = pokemon.get(current_key, {}).get("incomingEvolution")
        if not method:
            return None
        from_key = method.get("fromId")
        if not from_key or from_key not in family:
            return None
        methods.append(method)
        current_key = from_key
    return list(reversed(methods))


def timeline_candidate_from_record(pokemon: dict, target_key: str, source_key: str, record: dict, item_sources_by_key: dict[str, list[dict]]) -> dict | None:
    target = pokemon[target_key]
    family = target.get("family") or [target_key]
    path = evolution_path_between(pokemon, family, source_key, target_key)
    if path is None:
        return None

    checkpoints = [record_checkpoint(record)]
    requirements = []
    missing_items: list[str] = []
    for method in path:
        method_cp, item_requirements, method_missing = method_checkpoint(method, item_sources_by_key)
        checkpoints.append(method_cp)
        requirements.append({
            "fromId": method.get("fromId"),
            "fromName": method.get("fromName"),
            "label": method.get("label") or evolutionSummaryText(method),
            "requiredLevel": method_required_level(method),
            "items": item_requirements,
            "missingItems": method_missing,
            "specialGate": method.get("specialGate"),
        })
        missing_items.extend(method_missing)

    checkpoint = max_checkpoint(*checkpoints)
    via = pokemon.get(source_key, {}).get("displayName", record.get("pokemon", ""))
    if source_key == target_key:
        source = record.get("source", "Encounter")
        if record.get("startAvailable"):
            reason = "Wonder Trade entries are available from the start of the game."
        else:
            reason = f"Direct encounter by {checkpoint['label']}."
    else:
        source = "Evolution"
        reason = f"Evolves from {via} by {checkpoint['label']}."
    return {
        "checkpoint": checkpoint_payload(checkpoint),
        "source": source,
        "via": via,
        "baseId": source_key,
        "baseName": via,
        "encounter": {
            "pokemon": record.get("pokemon"),
            "area": record.get("area"),
            "method": record.get("method"),
            "level": record.get("level"),
            "minLevel": record.get("minLevel"),
            "rate": record.get("rate"),
            "source": record.get("source"),
            "checkpoint": checkpoint_payload(record_checkpoint(record)),
            "startAvailable": bool(record.get("startAvailable")),
        },
        "requirements": requirements,
        "missingItems": sorted(set(missing_items)),
        "startAvailable": bool(record.get("startAvailable")) and source_key == target_key,
        "reason": reason,
    }


def evolutionSummaryText(method: dict) -> str:
    return evolution_method_summary(method)


def attach_progression_timeline(pokemon: dict, item_sources: list[dict]) -> None:
    item_sources_by_key: dict[str, list[dict]] = defaultdict(list)
    for source in item_sources:
        item_sources_by_key[item_key(source["item"])].append(source)
    for sources in item_sources_by_key.values():
        sources.sort(key=lambda item: (item["checkpoint"]["sort"], item.get("gateLevel") or 0, item["source"], item["location"]))

    for key, mon in pokemon.items():
        candidates = []
        for family_key in mon.get("family", [key]):
            for record in pokemon.get(family_key, {}).get("directEncounters", []):
                candidate = timeline_candidate_from_record(pokemon, key, family_key, record, item_sources_by_key)
                if candidate:
                    candidates.append(candidate)

        if key in {"plusle", "minun"}:
            checkpoint = CHECKPOINT_BY_ID["first-stratum"]
            candidates.append({
                "checkpoint": checkpoint_payload(checkpoint),
                "source": "Story starter",
                "via": mon["displayName"],
                "baseId": key,
                "baseName": mon["displayName"],
                "encounter": {
                    "pokemon": mon["displayName"],
                    "area": "Starter",
                    "method": "Story starter",
                    "level": "",
                    "minLevel": None,
                    "rate": "",
                    "source": "Story starter",
                    "checkpoint": checkpoint_payload(checkpoint),
                },
                "requirements": [],
                "missingItems": [],
                "reason": "Nyx starts the journey with Plusle and Minun.",
            })

        candidates.sort(key=lambda candidate: (
            candidate["checkpoint"]["sort"],
            candidate["encounter"].get("minLevel") if candidate["encounter"].get("minLevel") is not None else 999,
            candidate["source"],
            candidate["via"],
        ))
        timeline = candidates[0] if candidates else {
            "checkpoint": checkpoint_payload(CHECKPOINT_BY_ID["unknown"]),
            "source": "Unknown",
            "via": "",
            "baseId": "",
            "baseName": "",
            "encounter": {},
            "requirements": [],
            "missingItems": [],
            "reason": "No parsed encounter or evolution route is available yet.",
        }
        mon["timeline"] = timeline
        checkpoint = timeline["checkpoint"]
        encounter = timeline.get("encounter", {})
        mon["availability"] = {
            "phase": checkpoint["label"],
            "sort": checkpoint["sort"],
            "source": timeline.get("source", ""),
            "details": f"{encounter.get('area', '')} - {encounter.get('method', '')}".strip(" -"),
            "via": timeline.get("via") or mon["displayName"],
            "level": encounter.get("level", ""),
            "rate": encounter.get("rate", ""),
            "checkpoint": checkpoint,
            "reason": timeline.get("reason", ""),
            "startAvailable": bool(timeline.get("startAvailable")),
        }


def resolve_pokemon_key(key: str, pokemon: dict) -> str:
    if key in pokemon:
        return key
    for suffix in ["-hisui", "-alola", "-galar"]:
        if key.endswith(suffix) and key[: -len(suffix)] in pokemon:
            return key[: -len(suffix)]
    compact_aliases = {
        "screamtail": "scream-tail",
        "sandyshock": "sandy-shock",
        "flut-mane": "flutter-mane",
        "ragingbolt": "raging-bolt",
        "gouginfire": "gouging-fire",
        "roar-moon": "roaring-moon",
        "walkinwake": "walking-wake",
    }
    return compact_aliases.get(key, "") if compact_aliases.get(key, "") in pokemon else ""


def pokemon_archetype(mon: dict) -> str:
    stats = mon.get("stats") or {}
    if not stats:
        return "Unknown"
    atk = stats.get("atk", 0)
    spa = stats.get("spa", 0)
    spe = stats.get("spe", 0)
    bulk = stats.get("hp", 0) + stats.get("def", 0) + stats.get("spd", 0)
    if bulk >= 310 and max(atk, spa) < 110:
        return "Tank"
    if spe >= 100 and max(atk, spa) >= 105:
        return "Fast Attacker"
    if atk >= 110 and spa >= 110:
        return "Mixed Attacker"
    if atk >= spa + 15:
        return "Physical Attacker"
    if spa >= atk + 15:
        return "Special Attacker"
    if bulk >= 280:
        return "Bulky Utility"
    return "Balanced"


def finalize_pokemon(pokemon: dict) -> list[dict]:
    rows = []
    for mon in pokemon.values():
        mon["roles"] = sorted(mon["roles"])
        mon["archetype"] = pokemon_archetype(mon)
        if mon["isVariant"] and "⭐" not in mon["displayName"]:
            mon["displayName"] = f"{mon['name']} ⭐"
        rows.append(mon)
    rows.sort(key=lambda mon: int(mon["gameDexId"] or 9999))
    return rows


def main() -> None:
    for path in [STATS_XLSX, WILD_XLSX, LEVELS_XLSX, TEAMBUILDING_PDF]:
        if not path.exists():
            raise FileNotFoundError(path)

    pokemon: dict[str, dict] = {}
    stats_wb = openpyxl.load_workbook(STATS_XLSX, data_only=True, read_only=False)
    parse_pokedex(stats_wb, pokemon)
    family_groups = parse_stats_and_learnsets(stats_wb, pokemon)
    sprite_count, unmatched_sprites = parse_etrian_variant_sprites(stats_wb, pokemon)
    odyssey_ability_definitions = parse_ability_definitions(stats_wb)
    standard_ability_definitions = fetch_standard_ability_definitions(
        ability_names_from_pokemon(pokemon),
        {entry["id"] for entry in odyssey_ability_definitions},
    )
    ability_definitions = merge_ability_definitions(odyssey_ability_definitions, standard_ability_definitions)

    wild_wb = openpyxl.load_workbook(WILD_XLSX, data_only=True, read_only=False)
    area_checkpoint_map = build_area_checkpoint_map(wild_wb)
    encounters = []
    encounters.extend(parse_wild_encounters(wild_wb, area_checkpoint_map))
    encounters.extend(parse_wonder_trade(wild_wb))
    encounters.extend(parse_naval_explorations(wild_wb))
    tms, tutors = parse_tms_and_tutors(wild_wb, area_checkpoint_map)

    levels_wb = openpyxl.load_workbook(LEVELS_XLSX, data_only=True, read_only=False)
    level_caps = parse_level_caps(levels_wb)
    sidequests = parse_sidequests(levels_wb, area_checkpoint_map)

    guide_entries = parse_teambuilding_guide(TEAMBUILDING_PDF)
    assign_families(pokemon, family_groups)
    assign_evolution_metadata(pokemon)
    attach_encounters(pokemon, encounters)
    attach_guide_entries(pokemon, guide_entries)
    evolution_item_names = evolution_item_names_from_pokemon(pokemon)
    evolution_item_sources = parse_evolution_item_sources(wild_wb, sidequests, evolution_item_names, area_checkpoint_map, encounters)
    attach_progression_timeline(pokemon, evolution_item_sources)

    roles = sorted({role for mon in pokemon.values() for role in mon.get("roles", [])})
    types = sorted({type_name for mon in pokemon.values() for type_name in mon.get("types", [])})

    payload = {
        "meta": {
            "title": "Pokemon Odyssey Team Planner",
            "version": "4.1.1",
            "generatedFrom": [
                STATS_XLSX.name,
                WILD_XLSX.name,
                LEVELS_XLSX.name,
                TEAMBUILDING_PDF.name,
            ],
            "sourceNotes": [
                "Availability uses the earliest parsed direct encounter for the Pokemon or its evolution family, then applies evolution level and item gates.",
                "Guide availability tags come from the teambuilding PDF as role context; encounter timing comes from parsed documentation and manual rules listed here.",
                "Exact story/boss details are kept in source records but summarized in the UI by default.",
                "Ability tooltips use Odyssey's new/buffed ability sheet first, with standard ability text cached from PokeAPI.",
                "Etrian Variant portrait art uses NORMAL embedded workbook sprites when a matching Pokemon exists.",
                "Timeline availability uses encounter order, stratum level caps, evolution levels, and parsed evolution-item sources.",
                "Manual timeline rule: Wonder Trade Pokemon are available from the start of the game.",
                "Manual timeline rule: Fire, Water, Thunder, and Sun Stones are available from the Talrega Lottery by the First Stratum cap.",
                "Napier's Shop treasure tiers use conservative gate-level assumptions from documented sea boss and naval F.O.E. levels.",
            ],
            "spriteAssets": sprite_count,
        },
        "progression": PROGRESSION_CHECKPOINTS,
        "pokemon": finalize_pokemon(pokemon),
        "encounters": sorted(encounters, key=sort_encounter),
        "guideEntries": guide_entries,
        "abilityDefinitions": ability_definitions,
        "evolutionItemSources": evolution_item_sources,
        "levelCaps": level_caps,
        "sidequests": sidequests,
        "tms": tms,
        "moveTutors": tutors,
        "facets": {
            "types": types,
            "roles": roles,
            "checkpoints": [checkpoint_payload(checkpoint) for checkpoint in PROGRESSION_CHECKPOINTS if checkpoint["id"] != "unknown"],
            "archetypes": sorted({pokemon_archetype(mon) for mon in pokemon.values()}),
        },
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT_JSON}")
    print(f"Pokemon: {len(payload['pokemon'])}")
    print(f"Encounters: {len(encounters)}")
    print(f"Guide entries: {len(guide_entries)}")
    print(f"Ability definitions: {len(ability_definitions)}")
    print(f"Evolution item sources: {len(evolution_item_sources)}")
    print(f"Etrian Variant sprites: {sprite_count}")
    if unmatched_sprites:
        print(f"Unmatched Etrian Variant sprites: {', '.join(unmatched_sprites)}")


if __name__ == "__main__":
    main()
